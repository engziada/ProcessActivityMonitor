"""
Utilities for exporting process monitoring data to Excel and PDF formats.
"""
import os
import configparser
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fpdf import FPDF
from models import Session, MonitoredProcess, ProcessActivityLog
from console_utils import format_time_delta

# Load configuration
config = configparser.ConfigParser()
config.read('config.ini')

# Get export directory from config
EXPORT_DIR = config['Export']['export_directory']

def ensure_export_dir():
    """Ensure the export directory exists."""
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)

def generate_filename(prefix, extension):
    """Generate a filename with timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(EXPORT_DIR, f"{prefix}_{timestamp}.{extension}")

def export_to_excel():
    """
    Export process monitoring data to Excel.

    Returns:
        str: Path to the created Excel file
    """
    ensure_export_dir()
    filename = generate_filename("process_activity", "xlsx")

    wb = Workbook()

    # Create Process Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Process Summary"

    # Create Activity Logs sheet
    ws_logs = wb.create_sheet("Activity Logs")

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    centered = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Process Summary headers
    headers_summary = ["Process Name", "PID", "Last Seen", "Total Uptime (HH:MM:SS)"]
    for col, header in enumerate(headers_summary, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = border

    # Activity Logs headers
    headers_logs = ["Process Name", "PID", "Start Time", "End Time", "Last Activity", "Duration (HH:MM:SS)"]
    for col, header in enumerate(headers_logs, 1):
        cell = ws_logs.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
        cell.border = border

    # Get data from database
    session = Session()
    try:
        # Process Summary data
        processes = session.query(MonitoredProcess).all()
        for row, proc in enumerate(processes, 2):
            ws_summary.cell(row=row, column=1, value=proc.process_name)
            ws_summary.cell(row=row, column=2, value=proc.pid)
            ws_summary.cell(row=row, column=3, value=proc.last_seen.strftime("%Y-%m-%d %H:%M:%S"))

            # Format uptime as HH:MM:SS
            if proc.last_uptime_seconds is not None:
                uptime_formatted = format_time_delta(proc.last_uptime_seconds)
                ws_summary.cell(row=row, column=4, value=uptime_formatted)
            else:
                ws_summary.cell(row=row, column=4, value="N/A")

            # Apply borders
            for col in range(1, 5):
                ws_summary.cell(row=row, column=col).border = border

        # Activity Logs data
        logs = session.query(ProcessActivityLog).join(MonitoredProcess).all()
        for row, log in enumerate(logs, 2):
            ws_logs.cell(row=row, column=1, value=log.process.process_name)
            ws_logs.cell(row=row, column=2, value=log.process.pid)
            ws_logs.cell(row=row, column=3, value=log.start_time.strftime("%Y-%m-%d %H:%M:%S"))

            end_time_str = log.end_time.strftime("%Y-%m-%d %H:%M:%S") if log.end_time else "Still Running"
            ws_logs.cell(row=row, column=4, value=end_time_str)

            ws_logs.cell(row=row, column=5, value=log.last_activity_time.strftime("%Y-%m-%d %H:%M:%S"))

            # Format duration as HH:MM:SS
            if log.session_uptime_seconds is not None:
                duration_formatted = format_time_delta(log.session_uptime_seconds)
                ws_logs.cell(row=row, column=6, value=duration_formatted)
            else:
                ws_logs.cell(row=row, column=6, value="N/A")

            # Apply borders
            for col in range(1, 7):
                ws_logs.cell(row=row, column=col).border = border

        # Auto-adjust column widths
        for sheet in [ws_summary, ws_logs]:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width

    finally:
        session.close()

    # Save workbook
    wb.save(filename)
    return filename

class PDF(FPDF):
    """Custom PDF class with header and footer."""

    def header(self):
        # Logo (if you have one)
        # self.image('logo.png', 10, 8, 33)

        # Title
        self.set_font('Arial', 'B', 15)
        self.cell(0, 10, 'Process Activity Monitor Report', 0, 1, 'C')

        # Date
        self.set_font('Arial', 'I', 10)
        self.cell(0, 10, f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 1, 'C')

        # Line break
        self.ln(10)

    def footer(self):
        # Position at 1.5 cm from bottom
        self.set_y(-15)

        # Footer text
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

def export_to_pdf():
    """
    Export process monitoring data to PDF.

    Returns:
        str: Path to the created PDF file
    """
    ensure_export_dir()
    filename = generate_filename("process_activity", "pdf")

    pdf = PDF()
    pdf.add_page()

    # Process Summary section
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, 'Monitored Processes Summary', 0, 1, 'L')
    pdf.ln(5)

    # Table header
    pdf.set_fill_color(221, 235, 247)  # Light blue
    pdf.set_font('Arial', 'B', 10)

    col_widths = [60, 20, 60, 50]
    headers = ["Process Name", "PID", "Last Seen", "Total Uptime (HH:MM:SS)"]

    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, 1, 0, 'C', 1)
    pdf.ln()

    # Table data
    pdf.set_font('Arial', '', 10)

    session = Session()
    try:
        # Process Summary data
        processes = session.query(MonitoredProcess).all()
        for proc in processes:
            pdf.cell(col_widths[0], 10, proc.process_name, 1, 0, 'L')
            pdf.cell(col_widths[1], 10, str(proc.pid), 1, 0, 'C')
            pdf.cell(col_widths[2], 10, proc.last_seen.strftime("%Y-%m-%d %H:%M:%S"), 1, 0, 'C')

            # Format uptime as HH:MM:SS
            uptime_text = "N/A"
            if proc.last_uptime_seconds is not None:
                uptime_text = format_time_delta(proc.last_uptime_seconds)
            pdf.cell(col_widths[3], 10, uptime_text, 1, 0, 'R')
            pdf.ln()

        # Activity Logs section
        pdf.ln(10)
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'Activity Logs', 0, 1, 'L')
        pdf.ln(5)

        # Table header
        pdf.set_fill_color(221, 235, 247)  # Light blue
        pdf.set_font('Arial', 'B', 10)

        log_col_widths = [40, 20, 40, 40, 40]
        log_headers = ["Process Name", "PID", "Start Time", "End Time", "Duration (HH:MM:SS)"]

        for i, header in enumerate(log_headers):
            pdf.cell(log_col_widths[i], 10, header, 1, 0, 'C', 1)
        pdf.ln()

        # Table data
        pdf.set_font('Arial', '', 10)

        logs = session.query(ProcessActivityLog).join(MonitoredProcess).all()
        for log in logs:
            # Check if we need a new page
            if pdf.get_y() > 250:
                pdf.add_page()

                # Reprint headers on new page
                pdf.set_fill_color(221, 235, 247)
                pdf.set_font('Arial', 'B', 10)
                for i, header in enumerate(log_headers):
                    pdf.cell(log_col_widths[i], 10, header, 1, 0, 'C', 1)
                pdf.ln()
                pdf.set_font('Arial', '', 10)

            # Handle None values for end_time
            end_time_str = "Still Running"
            if log.end_time is not None:
                end_time_str = log.end_time.strftime("%Y-%m-%d %H:%M:%S")

            # Format duration as HH:MM:SS
            duration_str = "N/A"
            if log.session_uptime_seconds is not None:
                duration_str = format_time_delta(log.session_uptime_seconds)

            pdf.cell(log_col_widths[0], 10, log.process.process_name, 1, 0, 'L')
            pdf.cell(log_col_widths[1], 10, str(log.process.pid), 1, 0, 'C')
            pdf.cell(log_col_widths[2], 10, log.start_time.strftime("%Y-%m-%d %H:%M:%S"), 1, 0, 'C')
            pdf.cell(log_col_widths[3], 10, end_time_str, 1, 0, 'C')
            pdf.cell(log_col_widths[4], 10, duration_str, 1, 0, 'R')
            pdf.ln()

    finally:
        session.close()

    # Save PDF
    pdf.output(filename)
    return filename
